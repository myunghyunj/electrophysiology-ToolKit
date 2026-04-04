from __future__ import annotations

import argparse
import csv
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from scipy.interpolate import PchipInterpolator

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

LIKELIHOOD_THRESH = 0.90
FPS_ORIGINAL = 30
FPS_TARGET = 3
RESAMPLE_FACTOR = FPS_ORIGINAL // FPS_TARGET


@dataclass
class TrackingResult:
    basename: str
    rows: list[dict[str, float]]
    total_distance_pixels: float
    total_time_sec: float
    avg_speed_pixels_per_sec: float
    max_speed_pixels_per_sec: float


@dataclass
class BinnedAnimal:
    filename: str
    distances_mm: np.ndarray


def discover_inputs(paths: Iterable[str], patterns: tuple[str, ...]) -> list[Path]:
    files: list[Path] = []
    for raw in paths:
        path = Path(raw)
        if path.is_dir():
            for pattern in patterns:
                files.extend(sorted(path.glob(pattern)))
        elif path.is_file():
            files.append(path)
    deduped: list[Path] = []
    seen = set()
    for file in files:
        resolved = file.resolve()
        if resolved not in seen:
            seen.add(resolved)
            deduped.append(file)
    return deduped


def _to_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return float('nan')


def format_bin_value(value: float) -> str:
    rounded = round(value)
    return str(int(rounded)) if abs(value - rounded) < 1e-9 else f'{value:g}'


def sem_from_values(matrix: np.ndarray) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    if matrix.size == 0:
        return np.array([]), np.array([]), np.array([])
    mean = np.nanmean(matrix, axis=0)
    counts = np.sum(~np.isnan(matrix), axis=0)
    sem = np.zeros(matrix.shape[1], dtype=float)
    valid = counts > 1
    if np.any(valid):
        sem[valid] = np.nanstd(matrix[:, valid], axis=0, ddof=1) / np.sqrt(counts[valid])
    return mean, sem, counts


def read_dlc_csv(file: Path) -> dict[str, np.ndarray]:
    with file.open(newline='') as f:
        rows = list(csv.reader(f))
    if len(rows) < 4:
        raise ValueError(f'DLC CSV is too short: {file}')
    bodyparts = rows[1]
    head_idx = [i for i, value in enumerate(bodyparts) if value.strip().lower() == 'head']
    if not head_idx:
        raise ValueError(f'No head marker found in {file.name}')
    x_col = head_idx[0]
    y_col = x_col + 1
    like_col = x_col + 2
    frames, head_x, head_y, likelihood = [], [], [], []
    for row in rows[3:]:
        if len(row) <= like_col:
            continue
        frames.append(_to_float(row[0]))
        head_x.append(_to_float(row[x_col]))
        head_y.append(_to_float(row[y_col]))
        likelihood.append(_to_float(row[like_col]))
    return {'frame': np.asarray(frames, dtype=float), 'head_x': np.asarray(head_x, dtype=float), 'head_y': np.asarray(head_y, dtype=float), 'likelihood': np.asarray(likelihood, dtype=float)}


def interpolate_series(values: np.ndarray) -> np.ndarray:
    x = np.arange(values.size)
    valid = np.isfinite(values)
    if valid.sum() == 0:
        return np.zeros_like(values)
    if valid.sum() == 1:
        return np.full_like(values, values[valid][0])
    return PchipInterpolator(x[valid], values[valid], extrapolate=True)(x)


def preprocess_file(file: Path, output_dir: Path) -> TrackingResult:
    data = read_dlc_csv(file)
    valid = data['likelihood'] >= LIKELIHOOD_THRESH
    x = data['head_x'].copy()
    y = data['head_y'].copy()
    x[~valid] = np.nan
    y[~valid] = np.nan
    x_interp = interpolate_series(x)
    y_interp = interpolate_series(y)
    sampled_idx = np.arange(0, x_interp.size, RESAMPLE_FACTOR)
    x_resampled = x_interp[sampled_idx]
    y_resampled = y_interp[sampled_idx]
    frames_resampled = data['frame'][sampled_idx]
    distances = np.sqrt(np.diff(x_resampled) ** 2 + np.diff(y_resampled) ** 2)
    time_sec = (frames_resampled - frames_resampled[0]) / FPS_ORIGINAL

    rows = []
    for idx, distance in enumerate(distances):
        rows.append({'Frame': float(frames_resampled[idx]), 'Time_sec': float(time_sec[idx]), 'Head_X': float(x_resampled[idx]), 'Head_Y': float(y_resampled[idx]), 'Distance_pixels': float(distance), 'Speed_pixels_per_sec': float(distance * FPS_TARGET)})

    basename = file.stem
    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / f'{basename}_head_tracking_3fps.csv', rows)
    write_xlsx_if_possible(output_dir / f'{basename}_head_tracking_3fps.xlsx', rows)
    save_head_plots(basename, x_resampled, y_resampled, rows, output_dir, file.name)
    total_distance = float(np.nansum(distances))
    total_time = float(time_sec[-1]) if time_sec.size else 0.0
    avg_speed = total_distance / total_time if total_time > 0 else 0.0
    max_speed = float(np.nanmax(distances) * FPS_TARGET) if distances.size else 0.0
    return TrackingResult(basename, rows, total_distance, total_time, avg_speed, max_speed)


def write_csv(path: Path, rows: list[dict[str, float]]) -> None:
    with path.open('w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['Frame', 'Time_sec', 'Head_X', 'Head_Y', 'Distance_pixels', 'Speed_pixels_per_sec'])
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_xlsx_if_possible(path: Path, rows: list[dict[str, float]]) -> None:
    if Workbook is None:
        return
    wb = Workbook()
    ws = wb.active
    headers = ['Frame', 'Time_sec', 'Head_X', 'Head_Y', 'Distance_pixels', 'Speed_pixels_per_sec']
    ws.append(headers)
    for row in rows:
        ws.append([row[key] for key in headers])
    wb.save(path)


def save_head_plots(basename: str, x_resampled: np.ndarray, y_resampled: np.ndarray, rows: list[dict[str, float]], output_dir: Path, original_name: str) -> None:
    color = '#3564e0' if 'SHAM' in original_name.upper() else '#d84a4a'
    fig, axes = plt.subplots(2, 1, figsize=(10, 8), constrained_layout=True)
    axes[0].plot(x_resampled, y_resampled, color=color, linewidth=1.5)
    axes[0].scatter(x_resampled[0], y_resampled[0], c='green', label='Start')
    axes[0].scatter(x_resampled[-1], y_resampled[-1], c='black', label='End')
    axes[0].set_xlabel('X (pixels)')
    axes[0].set_ylabel('Y (pixels)')
    axes[0].set_title(f'Head trajectory (3 fps) - {basename}')
    axes[0].axis('equal')
    axes[0].grid(True)
    axes[0].legend(loc='best', framealpha=1.0, fancybox=False)
    time_sec = np.array([row['Time_sec'] for row in rows], dtype=float)
    speed = np.array([row['Speed_pixels_per_sec'] for row in rows], dtype=float)
    axes[1].plot(time_sec, speed, color='black', linewidth=1)
    axes[1].set_xlabel('Time (seconds)')
    axes[1].set_ylabel('Speed (pixels/sec)')
    axes[1].set_title('Head speed over time')
    axes[1].grid(True)
    fig.savefig(output_dir / f'{basename}_head_analysis_3fps.png', dpi=200)
    fig.savefig(output_dir / f'{basename}_head_analysis_3fps.eps', format='eps')
    plt.close(fig)


def read_tracking_table(file: Path) -> list[dict[str, float]]:
    if file.suffix.lower() == '.csv':
        with file.open(newline='') as f:
            return [{key: _to_float(value) for key, value in row.items()} for row in csv.DictReader(f)]
    if file.suffix.lower() in {'.xlsx', '.xls'} and load_workbook is not None:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = [str(cell) for cell in rows[0]]
        return [{key: _to_float(value) for key, value in zip(header, row)} for row in rows[1:]]
    raise ValueError(f'Unsupported tracking table or missing openpyxl support: {file}')


def build_binned_animals(files: list[Path], bin_size_min: float, mm_per_pix: float) -> tuple[list[BinnedAnimal], int]:
    animals: list[BinnedAnimal] = []
    max_bins = 0
    for file in files:
        rows = read_tracking_table(file)
        time_min = np.array([row['Time_sec'] for row in rows], dtype=float) / 60.0
        dist_mm = np.array([row['Distance_pixels'] for row in rows], dtype=float) * mm_per_pix
        total_time = np.nanmax(time_min) if time_min.size else 0.0
        n_bins = max(1, int(math.ceil(total_time / bin_size_min)))
        binned = np.zeros(n_bins)
        for idx in range(n_bins):
            start = idx * bin_size_min
            end = (idx + 1) * bin_size_min
            mask = (time_min >= start) & (time_min < end)
            binned[idx] = float(np.nansum(dist_mm[mask])) if np.any(mask) else 0.0
        animals.append(BinnedAnimal(file.stem, binned))
        max_bins = max(max_bins, n_bins)
    return animals, max_bins


def padded_matrix(animals: list[BinnedAnimal], max_bins: int) -> np.ndarray:
    matrix = np.full((len(animals), max_bins), np.nan)
    for row_idx, animal in enumerate(animals):
        matrix[row_idx, :animal.distances_mm.size] = animal.distances_mm
    return matrix


def plot_group(ax, animals: list[BinnedAnimal], max_bins: int, bin_size_min: float, title: str, color: str) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    if not animals:
        ax.text(0.5, 0.5, f'No {title} data', ha='center', va='center', transform=ax.transAxes)
        ax.set_axis_off()
        return np.array([]), np.array([]), np.array([])
    matrix = padded_matrix(animals, max_bins)
    mean, sem, counts = sem_from_values(matrix)
    x = np.arange(1, max_bins + 1) * bin_size_min
    ax.bar(x, mean, color=color, edgecolor='black')
    ax.errorbar(x, mean, yerr=sem, fmt='k.', linewidth=1.2, capsize=3)
    for animal in animals:
        xs = x[:animal.distances_mm.size]
        ax.scatter(xs, animal.distances_mm, color='black', s=18)
    for idx, count in enumerate(counts):
        if count > 0:
            y = max(mean[idx] + sem[idx], 0) * 1.05
            ax.text(x[idx], y, f'n={int(count)}', ha='center', fontsize=8)
    ax.set_xlabel('Time (minutes)')
    ax.set_ylabel('Distance (mm)')
    ax.set_title(f'{title} Group (n={len(animals)})')
    ax.set_xticks(x)
    labels = [f'{format_bin_value((i) * bin_size_min)}-{format_bin_value((i + 1) * bin_size_min)}' for i in range(max_bins)]
    ax.set_xticklabels(labels, rotation=45)
    ax.grid(True)
    return mean, sem, counts


def save_summary_csv(path: Path, rows: list[dict[str, float | str]]) -> None:
    if not rows:
        return
    with path.open('w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def save_summary_xlsx_if_possible(path: Path, rows: list[dict[str, float | str]]) -> None:
    if Workbook is None or not rows:
        return
    wb = Workbook()
    ws = wb.active
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[key] for key in headers])
    wb.save(path)


def save_binned_outputs(sham: list[BinnedAnimal], stim: list[BinnedAnimal], max_bins: int, bin_size_min: float, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    fig, axes = plt.subplots(1, 2, figsize=(14, 6), constrained_layout=True)
    sham_mean, sham_sem, sham_n = plot_group(axes[0], sham, max_bins, bin_size_min, 'SHAM', '#4a7bd8')
    stim_mean, stim_sem, stim_n = plot_group(axes[1], stim, max_bins, bin_size_min, 'STIM', '#d84a4a')
    bin_label = format_bin_value(bin_size_min)
    fig.savefig(output_dir / f'binned_locomotion_{bin_label}min.png', dpi=200)
    fig.savefig(output_dir / f'binned_locomotion_{bin_label}min.eps', format='eps')
    plt.close(fig)
    summary_rows = []
    for idx in range(max_bins):
        row: dict[str, float | str] = {'Bin': f'{format_bin_value(idx * bin_size_min)}-{format_bin_value((idx + 1) * bin_size_min)} min'}
        if sham_mean.size:
            row['SHAM_Mean'] = float(sham_mean[idx])
            row['SHAM_SEM'] = float(sham_sem[idx])
            row['SHAM_N'] = int(sham_n[idx])
        if stim_mean.size:
            row['STIM_Mean'] = float(stim_mean[idx])
            row['STIM_SEM'] = float(stim_sem[idx])
            row['STIM_N'] = int(stim_n[idx])
        summary_rows.append(row)
    save_summary_csv(output_dir / f'binned_locomotion_summary_{bin_label}min.csv', summary_rows)
    save_summary_xlsx_if_possible(output_dir / f'binned_locomotion_summary_{bin_label}min.xlsx', summary_rows)


def sample_sem(values: np.ndarray) -> float:
    return float(np.std(values, ddof=1) / np.sqrt(values.size)) if values.size > 1 else 0.0


def print_group_summary(animals: list[BinnedAnimal], group_name: str, bin_size_min: float) -> None:
    if not animals:
        print(f'{group_name}: no data')
        return
    totals = np.array([np.sum(animal.distances_mm) for animal in animals], dtype=float)
    mean_bins = np.array([np.mean(animal.distances_mm) for animal in animals], dtype=float)
    print(f'{group_name}: animals={len(animals)}')
    print(f'  total distance mean±SEM = {totals.mean():.1f} ± {sample_sem(totals):.1f} mm')
    print(f'  total distance range = {totals.min():.1f} – {totals.max():.1f} mm')
    print(f'  average distance per {bin_size_min:g}-min bin = {mean_bins.mean():.1f} ± {sample_sem(mean_bins):.1f} mm')




def unique_tracking_files(files: list[Path]) -> list[Path]:
    preferred: dict[str, Path] = {}
    for file in sorted(files):
        key = file.stem
        if key.endswith('_head_tracking_3fps'):
            key = key[:-len('_head_tracking_3fps')]
        current = preferred.get(key)
        if current is None:
            preferred[key] = file
            continue
        if current.suffix.lower() != '.xlsx' and file.suffix.lower() == '.xlsx':
            preferred[key] = file
    return list(preferred.values())

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Python CLI for locomotion preprocessing and analysis')
    subparsers = parser.add_subparsers(dest='command', required=True)
    preprocess = subparsers.add_parser('preprocess', help='Convert DLC CSV exports to tracking tables')
    preprocess.add_argument('inputs', nargs='+', help='CSV files or directories')
    preprocess.add_argument('--output-dir', default='python_output', help='Output directory')
    analyze = subparsers.add_parser('analyze', help='Analyze binned locomotion from tracking outputs')
    analyze.add_argument('inputs', nargs='+', help='Tracking CSV/XLSX files or directories')
    analyze.add_argument('--output-dir', default='python_output', help='Output directory')
    analyze.add_argument('--bin-size-min', type=float, default=5.0, help='Bin size in minutes')
    analyze.add_argument('--mm-per-pix', type=float, required=True, help='Millimeters per pixel')
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.command == 'preprocess':
        files = discover_inputs(args.inputs, ('*.csv',))
        if not files:
            raise SystemExit('No DLC CSV files found.')
        output_dir = Path(args.output_dir)
        for file in files:
            result = preprocess_file(file, output_dir)
            print(f'{result.basename}: total distance={result.total_distance_pixels:.2f} px, total time={result.total_time_sec:.2f} s, avg speed={result.avg_speed_pixels_per_sec:.2f} px/s, max speed={result.max_speed_pixels_per_sec:.2f} px/s')
        return
    files = discover_inputs(args.inputs, ('*_head_tracking_3fps.csv', '*_head_tracking_3fps.xlsx', '*.csv', '*.xlsx'))
    files = unique_tracking_files(files)
    if not files:
        raise SystemExit('No tracking outputs found.')
    sham_files = [file for file in files if 'SHAM' in file.name.upper()]
    stim_files = [file for file in files if 'STIM' in file.name.upper()]
    sham, sham_max = build_binned_animals(sham_files, args.bin_size_min, args.mm_per_pix)
    stim, stim_max = build_binned_animals(stim_files, args.bin_size_min, args.mm_per_pix)
    max_bins = max(sham_max, stim_max, 1)
    output_dir = Path(args.output_dir)
    save_binned_outputs(sham, stim, max_bins, args.bin_size_min, output_dir)
    print_group_summary(sham, 'SHAM', args.bin_size_min)
    print_group_summary(stim, 'STIM', args.bin_size_min)
    print(f'Outputs written to {output_dir}')


if __name__ == '__main__':
    main()
